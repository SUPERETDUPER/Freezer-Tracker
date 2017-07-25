"""
MIT License

Copyright (c) 2017 Martin Staadecker

Permission is hereby granted, free of charge, to any person obtaining a copy
of this software and associated documentation files (the "Software"), to deal
in the Software without restriction, including without limitation the rights
to use, copy, modify, merge, publish, distribute, sublicense, and/or sell
copies of the Software, and to permit persons to whom the Software is
furnished to do so, subject to the following conditions:

The above copyright notice and this permission notice shall be included in all
copies or substantial portions of the Software.

THE SOFTWARE IS PROVIDED "AS IS", WITHOUT WARRANTY OF ANY KIND, EXPRESS OR
IMPLIED, INCLUDING BUT NOT LIMITED TO THE WARRANTIES OF MERCHANTABILITY,
FITNESS FOR A PARTICULAR PURPOSE AND NONINFRINGEMENT. IN NO EVENT SHALL THE
AUTHORS OR COPYRIGHT HOLDERS BE LIABLE FOR ANY CLAIM, DAMAGES OR OTHER
LIABILITY, WHETHER IN AN ACTION OF CONTRACT, TORT OR OTHERWISE, ARISING FROM,
OUT OF OR IN CONNECTION WITH THE SOFTWARE OR THE USE OR OTHER DEALINGS IN THE
SOFTWARE.

Backend code with database operations.

"""
import os.path
import sys

import openpyxl

import config
import global_var
import helper

local_path_w_extension = config.local_path + global_var.db_extension


class Database:
    next_id = 10000

    def __init__(self):
        if os.path.isfile(local_path_w_extension):
            self.workbook = openpyxl.load_workbook(filename=local_path_w_extension)  # If file exists, load it
            self.ws = self.workbook.active
        else:
            self.workbook = openpyxl.Workbook()  # Else create it and create headers
            self.ws = self.workbook.active

            self.create_header()

        for index, row in enumerate(self.ws.iter_rows(row_offset=1)):  # Loop through every row

            empty = True

            for cell in row:
                if cell.value != "" and cell.value is not None:
                    empty = False

            if empty:  # If empty tracker is still True after looping through all the row's cells, this line is empty
                self.lastRow = index + 1
                break

            if self.next_id < row[
                global_var.columns[global_var.idColumn]].value + 1:  # Update to get latest id assigned
                self.next_id = row[global_var.columns[global_var.idColumn]].value + 1

    def create_header(self):  # Used to create a new database's headers
        for column in global_var.columns.keys():
            header_cell = self.ws.cell(row=1, column=global_var.columns[column] + 1)
            header_cell.value = column

    def save(self):  # Save workbook
        self.workbook.save(filename=local_path_w_extension)

    def add_item(self, row):  # Add a row to the database and return generated id
        empty_row = self.ws[self.lastRow + 1]  # Empty row to fill

        for index, cell in enumerate(global_var.columns):  # For every cell in empty_row

            value = row.get_row()[index]  # Get value to fill

            if value is not None:
                empty_row[index].value = row.get_row()[index]  # If value exists fill it in

            if index == 0:  # Generate index
                empty_row[index].value = self.next_id  # Assign batch number
                self.next_id += 1

        empty_row[global_var.columns[global_var.removedColumn]].value = "False"  # Set removed to false

        self.lastRow += 1

        self.save()
        upload()

        return self.next_id - 1

    def remove_item(self, batch_number):
        row = self.get_row(batch_number)  # Get row to set to removed
        if row == global_var.ERROR_NO_SUCH_ITEM:
            return global_var.ERROR_NO_SUCH_ITEM

        if row[global_var.columns[global_var.removedColumn]].value == "True":
            return global_var.ERROR_ITEM_REMOVED

        row[global_var.columns[global_var.removedColumn]].value = "True"

        row[global_var.columns[global_var.removedTimeColumn]].value = helper.get_current_date()

        self.save()
        upload()
        return True

    def get_info(self, batch_number):

        row = self.get_row(batch_number)  # get row with info

        if row == global_var.ERROR_NO_SUCH_ITEM:  # If row does not exist return False
            return global_var.ERROR_NO_SUCH_ITEM

        if row[global_var.columns[global_var.removedColumn]].value == "True":
            return global_var.ERROR_ITEM_REMOVED  # If row already removed return False

        return Row(batch_number=row[global_var.columns[global_var.idColumn]].value,
                   category=row[global_var.columns[global_var.typeColumn]].value,
                   subcategory=row[global_var.columns[global_var.subtypeColumn]].value,
                   weight=row[global_var.columns[global_var.weightColumn]].value,
                   entry_date=row[global_var.columns[global_var.timeColumn]].value)

    def get_row(self, batch_number):
        for index, row in enumerate(self.ws.iter_rows(row_offset=1)):  # Loop through every row
            if row[global_var.columns[global_var.idColumn]].value == batch_number:  # If batch number matches return row
                return row
        return global_var.ERROR_NO_SUCH_ITEM  # Else return -1


class Row:  # Row object storing row data
    def __init__(self, category=None, subcategory=None, weight=None, batch_number=None, entry_date=None):
        if entry_date is None:
            entry_date = helper.get_current_date()

        self.row = [None] * len(global_var.columns)

        self.row[global_var.columns[global_var.idColumn]] = batch_number
        self.row[global_var.columns[global_var.timeColumn]] = entry_date
        self.row[global_var.columns[global_var.typeColumn]] = category
        self.row[global_var.columns[global_var.subtypeColumn]] = subcategory
        self.row[global_var.columns[global_var.weightColumn]] = weight

    def get_row(self):
        return self.row

    def get_item(self, index):
        return self.row[index]


def upload():
    if sys.platform == "win32":
        os.system("copy " + local_path_w_extension + " " + config.upload_path + global_var.db_extension)
    else:
        os.system("cp ./" + local_path_w_extension + " " + config.upload_path + global_var.db_extension)


global_var.database = Database()
